import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.compat import range
from openpyxl.utils import FORMULAE
import openpyxl.drawing.image
import datetime


def load_from_file():
    """从已有表格文件加载工作簿"""
    wb = openpyxl.load_workbook('example.xlsx', guess_types=True, data_only=True, keep_vba=True)
    print(wb)


def get_workbook_attributes():
    """
    获取工作簿的属性：
        工作簿里工作表的名称
        根据名称获取工作表
        获取活跃的工作表
    """
    wb = openpyxl.load_workbook('example.xlsx')
    print(wb.sheetnames)
    sheet1 = wb['Sheet1']
    sheet2 = wb.active
    print(sheet1)
    print(sheet2)


def get_sheet_attributes():
    """名称，最大行和最大列，行生成器和列生成器"""
    wb = openpyxl.load_workbook('example.xlsx')
    sheet1 = wb.active
    print(sheet1)
    print(sheet1.title)
    print(sheet1.max_row)
    print(sheet1.max_column)
    print(sheet1.rows)
    print(sheet1.columns)


def read_from_cell():
    """读取单元格数据"""
    wb = openpyxl.load_workbook('example.xlsx')
    sheet1 = wb.active
    print(sheet1['A1'].value)
    print(sheet1.cell(row=1, column=2).value)


def get_rows_columns():
    """行/列生成器，每一行/列为一个元组，可迭代"""
    wb = openpyxl.load_workbook('example.xlsx')
    sheet1 = wb.active
    for cols in sheet1.iter_cols(min_row=1, min_col=1, max_col=3, max_row=2):
        for cell in cols:
            print(cell)
    for rows in sheet1.iter_rows(min_row=1, min_col=1, max_col=3, max_row=2):
        for cell in rows:
            print(cell)
    for cell in sheet1.rows:
        print(type(cell))
        print(cell)
    print(type(sheet1.rows))
    print(tuple(sheet1.columns))


def create_cells():
    """工作表创建时，单元格只有一个，当调用其他单元格时，再自动生成，即使不赋值"""
    wb2 = openpyxl.Workbook()
    ws1 = wb2.create_sheet('MySheet1', 0)
    ws2 = wb2.create_sheet('MySheet2', 1)
    for i in range(1, 101):
        ws1.cell(row=i, column=i)
        ws2.cell(row=i, column=i)
    print(ws1.max_row)
    print(ws1.max_column)
    wb2.save('balance.xlsx')


def save2template():
    """模板"""
    wb3 = openpyxl.Workbook()
    wb3.template = True
    wb3.save('document_template.xltx')
    # wb4 = openpyxl.load_workbook('document_template.xltx')
    # wb4.save('document.xlsx', as_template=False)


def change_column_number2letter():
    """列数字转化为字母"""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = 'range names'
    for row in range(1, 40):
        print(row)
        ws1.append(range(600))
    ws3 = wb.create_sheet(title='PI')
    ws3['F5'] = 3.14
    ws4 = wb.create_sheet(title='Data')
    for row in range(10, 20):
        for col in range(27, 54):
            ws4.cell(row=row, column=col, value='{}'.format(get_column_letter(col)))
    print(ws4['AA10'].value)
    wb.save('empty_book.xlsx')


def number_format():
    # 不会保存图标、图片
    wb6 = openpyxl.load_workbook('empty_book.xlsx', guess_types=True, data_only=True, keep_vba=True)
    wb6.close()
    wb6 = openpyxl.Workbook()
    ws = wb6.active
    ws['A1'] = datetime.datetime(2010, 7, 21)
    print(ws['A1'].value)
    print(ws['A1'].number_format)
    wb6.guess_types = True
    ws['B1'] = '3.14%'
    print(ws['B1'].value)
    print(ws['B1'].number_format)
    wb6.guess_types = False
    ws['B1'] = '3.14%'
    print(ws['B1'].value)


def insert_formula():
    wb7 = openpyxl.Workbook()
    ws = wb7.active
    ws['A1'] = '=SUM(1, 1)'
    wb7.save('formula.xlsx')
    print("HEX2DEC" in FORMULAE)


def merge_cells():
    wb7 = openpyxl.Workbook()
    ws = wb7.active
    ws.merge_cells('A2:D4')
    ws.unmerge_cells('A2:D4')
    ws.merge_cells(start_row=5, start_column=5, end_row=7, end_column=7)
    ws.unmerge_cells(start_row=5, start_column=5, end_row=7, end_column=7)
    wb7.save('merge.xlsx')


def insert_image():
    wb8 = openpyxl.Workbook()
    ws = wb8.active
    ws['A1'] = 'You should see three logos below'
    img = openpyxl.drawing.image.Image('sg.png')
    ws.add_image(img, 'A1')
    wb8.save('image.xlsx')


def fold_columns():
    pass


if __name__ == '__main__':
    pass
